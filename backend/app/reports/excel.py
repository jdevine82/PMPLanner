import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL   = PatternFill("solid", fgColor="2563EB")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
REFUSED_FONT  = Font(color="DC2626", bold=True)
SECTION_FONT  = Font(bold=True, size=12)
NORMAL_FONT   = Font(size=10)


def _header_row(ws, values: list[str], row: int) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def generate_excel(report_data: dict[str, Any]) -> bytes:
    wb = Workbook()
    customer = report_data["customer"]
    title = f"PM Report — {customer['company_name']}"

    # ── Sheet A: Asset Inventory ─────────────────────────────────────────────
    ws_a = wb.active
    ws_a.title = "Asset Inventory"
    ws_a["A1"] = title
    ws_a["A1"].font = SECTION_FONT
    ws_a["A2"] = f"Generated: {report_data['generated_date']}"
    _header_row(ws_a, ["Asset Name", "Serial Number", "Model Number", "Location"], 4)
    for i, row in enumerate(report_data["asset_inventory"], 5):
        ws_a.cell(i, 1, row["asset_name"])
        ws_a.cell(i, 2, row["serial_number"])
        ws_a.cell(i, 3, row["model_number"])
        ws_a.cell(i, 4, row["location"])
    _auto_width(ws_a)

    # ── Sheet B: Scheduling ──────────────────────────────────────────────────
    ws_b = wb.create_sheet("Scheduling")
    _header_row(ws_b, ["Asset", "Service Template", "Frequency", "Est. Hours", "Next Due", "Last Done"], 1)
    for i, row in enumerate(report_data["scheduling"], 2):
        ws_b.cell(i, 1, row["asset_name"])
        ws_b.cell(i, 2, row["service_title"])
        ws_b.cell(i, 3, row["frequency"])
        ws_b.cell(i, 4, row["estimated_hours"])
        ws_b.cell(i, 5, row["date_next_due"])
        ws_b.cell(i, 6, row["date_last_done"])
    _auto_width(ws_b)

    # ── Sheet C: Service History ─────────────────────────────────────────────
    ws_c = wb.create_sheet("Service History")
    _header_row(ws_c, ["Month", "Asset", "Service", "Status", "Actual Hours", "Refusal Reason"], 1)
    for i, row in enumerate(report_data["history"], 2):
        ws_c.cell(i, 1, row["month"])
        ws_c.cell(i, 2, row["asset_name"])
        ws_c.cell(i, 3, row["service_title"])
        status_cell = ws_c.cell(i, 4, row["status"])
        if row["status"] == "Refused by Customer":
            status_cell.font = REFUSED_FONT
        ws_c.cell(i, 5, row["actual_hours"])
        ws_c.cell(i, 6, row["refusal_reason"] or "")
    _auto_width(ws_c)

    # ── Sheet D: Forecast ────────────────────────────────────────────────────
    ws_d = wb.create_sheet(f"Forecast ({report_data['forecast_months']}mo)")
    _header_row(ws_d, ["Due Date", "Asset", "Service Template", "Est. Hours"], 1)
    for i, row in enumerate(report_data["forecast"], 2):
        ws_d.cell(i, 1, row["due_date"])
        ws_d.cell(i, 2, row["asset_name"])
        ws_d.cell(i, 3, row["service_title"])
        ws_d.cell(i, 4, row["estimated_hours"])
    _auto_width(ws_d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
